# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html

import openpyxl
import datetime
import os.path

class TwitchtrendPipeline(object):
    def process_item(self, item, spider):
        return item

class ExcelPipeline(object):
    def __init__(self):
        dt = datetime.date.today().strftime('%m%d%y')
        self.dbname = dt + '.xlsx'
        self.dbdir = os.path.dirname(os.path.realpath(__file__)) + '\\..\\database\\'
        if os.path.isfile(self.dbdir + self.dbname):
            self.wb = openpyxl.load_workbook(self.dbdir + self.dbname)
        else:
            self.wb = openpyxl.Workbook()
            ws = self.wb.active
            ws['H1'] = 'NextLine:'
            ws['I1'] = 2
            ws['A1'] = 'Date'
            ws['B1'] = 'Time'
            ws['C1'] = 'Rank'
            ws['D1'] = 'Games'
            ws['E1'] = 'Channels'
            ws['F1'] = 'Viewers'
            ws.column_dimensions['A'].width = 11
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 5
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.freeze_panes = 'A2'


    def process_item(self, item, spider):
        ws = self.wb.active
        next_input_line = ws['I1'].value
        items_number = len(item['rank'])
        for n in item['rank']:
            ws['A' + str(int(n) - 1 + next_input_line)] = item['date']
            ws['B' + str(int(n) - 1 + next_input_line)] = item['time']
            ws['C' + str(int(n) - 1 + next_input_line)] = int(item['rank'][int(n) - 1])
            ws['D' + str(int(n) - 1 + next_input_line)] = item['games'][int(n) - 1]
            ws['E' + str(int(n) - 1 + next_input_line)] = int(item['channels'][int(n) - 1])
            ws['F' + str(int(n) - 1 + next_input_line)] = int(item['viewers'][int(n) - 1])

        next_input_line = next_input_line + items_number
        ws['I1'] = next_input_line
        return item

    def close_spider(self, spider):
        dirName = self.dbdir + self.dbname
        self.wb.save(dirName)


